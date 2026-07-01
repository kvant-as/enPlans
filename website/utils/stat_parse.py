from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

from openpyxl import load_workbook


class StatParseError(Exception):
    pass


@dataclass
class StatCellValue:
    row_code: str
    row_name: str
    column_code: str
    value: float


@dataclass
class ParsedStatPlan:
    type: str
    year: int
    org_name_in_file: str
    okpo_from_filename: Optional[str]
    source_filename: str
    values: List[StatCellValue] = field(default_factory=list)


FILENAME_RE = re.compile(r"^(?P<okpo>\d{6,10})_(?P<year>\d{4})_(?P<form>4|12)\.xlsx?$", re.IGNORECASE)

PERIOD_RE = re.compile(
    r"за\s+(?P<period>[A-Za-zА-Яа-яЁё\-]+(?:\s*-\s*[A-Za-zА-Яа-яЁё]+)?)\s+(?P<year>20\d{2})\s*год",
    re.IGNORECASE,
)

ORG_NAME_RE = re.compile(r'["«](?P<name>[^"»]+)["»"]')

ROWS_12TEK = {110, 111, 112, 113, 120, 130, 140, 141, 142, 143, 150, 260}

COLS_12TEK_YTD = {1: 3, 2: 4, 3: 5, 4: 6, 5: 7}


def _col_4tek(graph: int) -> int:
    return graph + 3


def _first_n_rows_text(ws, n: int = 10) -> str:
    chunks = []
    for r in range(1, n + 1):
        v = ws.cell(row=r, column=1).value
        if v:
            chunks.append(str(v))
    return "\n".join(chunks)


def extract_okpo_from_filename(filename: str) -> Optional[str]:
    m = FILENAME_RE.match(filename)
    return m.group("okpo") if m else None


def extract_year_from_filename(filename: str) -> Optional[int]:
    m = FILENAME_RE.match(filename)
    return int(m.group("year")) if m else None


def detect_report_type(ws) -> str:
    head = _first_n_rows_text(ws, 10)
    if re.search(r"4\s*-\s*ТЭК", head, re.IGNORECASE):
        return "4-tek"
    if re.search(r"12\s*-\s*ТЭК", head, re.IGNORECASE):
        return "12-tek"
    raise StatParseError(
        'Не удалось определить тип отчёта по содержимому файла '
        '(в первых строках не найдено "12-ТЭК" или "4-ТЭК")'
    )


def parse_org_name(ws) -> str:
    for r in range(1, 7):
        v = ws.cell(row=r, column=1).value
        if not v:
            continue
        m = ORG_NAME_RE.search(str(v))
        if m:
            return m.group("name").strip()
    raise StatParseError("Не найдено название организации в файле (ожидалось в кавычках, напр. ОАО \"...\")")


def _to_float(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".").replace("\xa0", "")
    if s in ("", "-", "х", "x", "Х", "X"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_12tek(ws) -> List[StatCellValue]:
    values: List[StatCellValue] = []
    for row in ws.iter_rows():
        code_cell = row[0].value
        if not isinstance(code_cell, (int, float)):
            continue
        code_int = int(code_cell)
        if code_int not in ROWS_12TEK:
            continue

        r = row[0].row
        row_code = str(code_int)
        row_name = ws.cell(row=r, column=2).value or ""

        for graph, col in COLS_12TEK_YTD.items():
            val = _to_float(ws.cell(row=r, column=col).value)
            if val is not None:
                values.append(StatCellValue(row_code, row_name, str(graph), val))

    return values


def parse_4tek(ws) -> List[StatCellValue]:
    values: List[StatCellValue] = []
    for row in ws.iter_rows():
        code_cell = row[0].value
        if not isinstance(code_cell, (int, float)):
            continue

        r = row[0].row
        row_code = str(int(code_cell))
        row_name = ws.cell(row=r, column=2).value or ""

        for graph in range(1, 13):
            col = _col_4tek(graph)
            val = _to_float(ws.cell(row=r, column=col).value)
            if val is not None:
                values.append(StatCellValue(row_code, row_name, str(graph), val))

    return values


def parse_stat_file(file_path: str, filename: Optional[str] = None) -> ParsedStatPlan:
    filename = filename or file_path.rsplit("/", 1)[-1]

    wb = load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    report_type = detect_report_type(ws)
    org_name = parse_org_name(ws)
    okpo = extract_okpo_from_filename(filename)
    year = extract_year_from_filename(filename) or 0

    if report_type == "12-tek":
        values = parse_12tek(ws)
    else:
        values = parse_4tek(ws)

    if not values:
        raise StatParseError("В файле не найдено ни одной распознанной строки данных")

    return ParsedStatPlan(
        type=report_type,
        year=year,
        org_name_in_file=org_name,
        okpo_from_filename=okpo,
        source_filename=filename,
        values=values,
    )