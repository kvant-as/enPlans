import io
import os
import zipfile
from flask import current_app
from .models import Direction, Indicator, Plan, Event, IndicatorUsage
from sqlalchemy.orm import sessionmaker, joinedload

export_tasks = {}

def type_of_export(plan: Plan) -> str:
    if plan.org_id:
        for indicator_usage in plan.indicators_usage:
            if indicator_usage.indicator.code == '260' and indicator_usage.QYearCurrent is not None:
                try:
                    q_year_next = float(indicator_usage.QYearCurrent)
                    if q_year_next >= 25000:
                        return "org_large"
                    else:
                        return "org_small"
                except (ValueError, TypeError):
                    return "org_small"
        return "org_small"
    
    # if plan.ministry_id:
    #     return "ministry"
    
    # if plan.region_id:
    #     return "region"
    
    raise ValueError("Error for read type of plan")

def create_export_archive_async(export_format, task_id, user_id, plan_ids, app):
    with app.app_context():
        from website import db
        Session = sessionmaker(bind=db.engine)
        session = Session()
        
        try:
            export_tasks[task_id] = {
                'status': 'processing',
                'progress': 0,
                'format': export_format,
                'user_id': user_id
            }
            
            temp_dir = os.path.join(current_app.config.get('UPLOAD_FOLDER', '/tmp'), 'exports', task_id)
            os.makedirs(temp_dir, exist_ok=True)
            
            zip_path = os.path.join(temp_dir, f'plans_export_{task_id}.zip')
            export_tasks[task_id]['progress'] = 10
            
            total_plans = len(plan_ids)
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for idx, plan_id in enumerate(plan_ids):
                    progress = 10 + int((idx + 1) / total_plans * 80)
                    export_tasks[task_id]['progress'] = progress
                    
                    plan = session.query(Plan).options(
                        joinedload(Plan.indicators_usage).joinedload(IndicatorUsage.indicator).joinedload(Indicator.unit),
                        joinedload(Plan.events).joinedload(Event.direction).joinedload(Direction.unit),
                        joinedload(Plan.organization),
                    ).filter(Plan.id == plan_id).first()
                    
                    if not plan:
                        continue
                    
                    if export_format == "xlsx":
                        file_stream, mime, filename = export_xlsx_single(plan)
                    elif export_format == "xml":
                        file_stream, mime, filename = export_xml_single(plan)
                    else:
                        continue
                    
                    zip_file.writestr(filename, file_stream.getvalue())
            
            export_tasks[task_id]['status'] = 'completed'
            export_tasks[task_id]['progress'] = 100
            export_tasks[task_id]['file_path'] = zip_path
            
        except Exception as e:
            current_app.logger.error(f"Export error for task {task_id}: {str(e)}", exc_info=True)
            export_tasks[task_id] = {
                'status': 'error',
                'error': str(e),
                'format': export_format,
                'user_id': user_id
            }
        finally:
            session.close()

def export_xml_single(plan: Plan):
        pass

def export_pdf_single(plan: Plan):
        pass

def export_xlsx_single(plan: Plan):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    export_type = type_of_export(plan)
    # ===============================
    # Шрифт и выравнивание
    # ===============================
    regular_font_9 = Font(name="Times New Roman", size=9)
    regular_font_9_italic = Font(name="Times New Roman", size=9, italic=True)
    regular_font_10 = Font(name="Times New Roman", size=10)
    regular_font_10_italic = Font(name="Times New Roman", size=10, italic=True)
    
    regular_font_11 = Font(name="Times New Roman", size=11)
    regular_font_11_italic = Font(name="Times New Roman", size=11, italic=True)
    
    regular_font_13 = Font(name="Times New Roman", size=13)
    bold_font_10 = Font(name="Times New Roman", size=10, bold=True)
    bold_font_11 = Font(name="Times New Roman", size=11, bold=True)
    bold_font_13 = Font(name="Times New Roman", size=13, bold=True)
    
    vertical_text = Alignment(horizontal="center", vertical="bottom", textRotation=90, wrap_text=True)
    top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    bottom = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    bottom_left = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    thin_bottom = Side(border_style="thin", color="000000")
    bottom_border = Border(bottom=thin_bottom)

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    def set_cell(ws, row_start, col_start, row_end=None, col_end=None, text="", 
                        font=regular_font_10, row_height=None, alignment=left,
                        merge_direction='both'):
        if merge_direction == 'horizontal':
            row_end = row_start
            if col_end is None:
                col_end = col_start
        elif merge_direction == 'vertical':
            col_end = col_start
            if row_end is None:
                row_end = row_start
        elif merge_direction == 'both':
            if row_end is None:
                row_end = row_start
            if col_end is None:
                col_end = col_start
        else:
            row_end = row_start
            col_end = col_start
        
        if row_start != row_end or col_start != col_end:
            ws.merge_cells(start_row=row_start, start_column=col_start, 
                        end_row=row_end, end_column=col_end)
        
        cell = ws.cell(row=row_start, column=col_start)
        cell.value = text
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        
        if row_height is not None:
            for row in range(row_start, row_end + 1):
                ws.row_dimensions[row].height = row_height
        
        return cell
    
    def page_settings(ws, print_area):
        ws.print_area = print_area      
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

    def get_org_name_by_okpo(okpo):
        dop_org_data = [
            ('Брестское областное управление', '100000001000'),
            ('Витебское областное управление', '200000002000'),
            ('Гомельское областное управление', '300000003000'),
            ('Гродненское областное управление', '400000004000'),
            ('Управление г. Минск', '500000005000'),
            ('Минское областное управление', '600000006000'),
            ('Могилевское областное управление', '700000007000'),
            ('Департамент по энергоэффективности', '800000008000'),
        ]
    
        if not okpo or len(str(okpo)) < 4:
            return ""  
        
        okpo_str = str(okpo)
        if len(okpo_str) >= 4:
            fourth_from_end = okpo_str[-4]  
        else:
            return ""
        
        for name, code in dop_org_data:
            if code.endswith(str(fourth_from_end) + "000"):
                return name
        
        return ""

    def org_small_title_xlsx(wb, plan):
        ws = wb.create_sheet("Титульный лист", 0)
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center
                    
            ws.merge_cells("B4:F4")
            ws["B4"].value = "_______________"
            ws["B4"].font = regular_font_11
            ws["B4"].alignment = left
                            
            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"{org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"областное (городское) управление по надзору за рациональным использованием ТЭР"
                        
            ws.merge_cells("B5:F5")
            ws["B5"].value = org_text
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        title_first_sign()
        title_second_sign()
        
        ws.merge_cells("A14:O15")
        ws["A14"].value = "ПЛАН МЕРОПРИЯТИЙ ПО ЭНЕРГОСБЕРЕЖЕНИЮ"
        ws["A14"].font = bold_font_13
        ws["A14"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{plan.organization.name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(наименование юридического лица)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"на {plan.year} год".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Целевые показатели:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:H25")
        ws["E25"].value = "энергосбережения"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:H26")
        ws["E26"].value = "по экономии ТЭР"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "по доле местных ТЭР в КПТ"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws.merge_cells("E28:H28")
        ws["E28"].value = "по доле ВИЭ в КПТ"
        ws["E28"].font = bold_font_11
        ws["E28"].alignment = left
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    
        ws["I28"].value = "-"
        ws["I28"].font = bold_font_11
        ws["I28"].alignment = center

        ws["J25"].value = f"{plan.energy_saving}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = center        
        ws["J26"].value = f"{plan.share_fuel}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = center        
        ws["J27"].value = f"{plan.saving_fuel}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = center        
        ws["J28"].value = f"{plan.share_energy}"
        ws["J28"].font = bold_font_11
        ws["J28"].alignment = center       

        ws["K25"].value = "%"
        ws["K25"].font = bold_font_11
        ws["K25"].alignment = center        
        ws["K26"].value = "т у.т."
        ws["K26"].font = bold_font_11
        ws["K26"].alignment = center        
        ws["K27"].value = "%"
        ws["K27"].font = bold_font_11
        ws["K27"].alignment = center        
        ws["K28"].value = "%"
        ws["K28"].font = bold_font_11
        ws["K28"].alignment = center
      
        page_settings(ws, print_area = "A1:O32")
        
        return ws
 
    def org_large_title_xlsx(wb, plan):
        ws = wb.create_sheet("Титульный лист", 0)
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center
                    
            ws.merge_cells("B4:F4")
            ws["B4"].value = "Департамента по энергоэффективности"
            ws["B4"].font = regular_font_11
            ws["B4"].alignment = left
                                
            ws.merge_cells("B5:F5")
            ws["B5"].value = 'Госстандарта'
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        title_first_sign()
        title_second_sign()
        
        ws.merge_cells("A14:O15")
        ws["A14"].value = "ПЛАН МЕРОПРИЯТИЙ ПО ЭНЕРГОСБЕРЕЖЕНИЮ"
        ws["A14"].font = bold_font_13
        ws["A14"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{plan.organization.name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(наименование юридического лица)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"на {plan.year} год".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Целевые показатели:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:H25")
        ws["E25"].value = "энергосбережения"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:H26")
        ws["E26"].value = "по экономии ТЭР"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "по доле местных ТЭР в КПТ"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws.merge_cells("E28:H28")
        ws["E28"].value = "по доле ВИЭ в КПТ"
        ws["E28"].font = bold_font_11
        ws["E28"].alignment = left
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    
        ws["I28"].value = "-"
        ws["I28"].font = bold_font_11
        ws["I28"].alignment = center

        ws["J25"].value = f"{plan.energy_saving}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = center        
        ws["J26"].value = f"{plan.share_fuel}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = center        
        ws["J27"].value = f"{plan.saving_fuel}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = center        
        ws["J28"].value = f"{plan.share_energy}"
        ws["J28"].font = bold_font_11
        ws["J28"].alignment = center       

        ws["K25"].value = "%"
        ws["K25"].font = bold_font_11
        ws["K25"].alignment = center        
        ws["K26"].value = "т у.т."
        ws["K26"].font = bold_font_11
        ws["K26"].alignment = center        
        ws["K27"].value = "%"
        ws["K27"].font = bold_font_11
        ws["K27"].alignment = center        
        ws["K28"].value = "%"
        ws["K28"].font = bold_font_11
        ws["K28"].alignment = center
                     
        page_settings(ws, print_area = "A1:O32")
        
        return ws
 
    def min_title_xlsx(wb, plan):
        ws = wb.create_sheet("Титульный лист", 0)
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center
                    
            ws.merge_cells("B4:F4")
            ws["B4"].value = "Департамента по энергоэффективности"
            ws["B4"].font = regular_font_11
            ws["B4"].alignment = left
                            
            ws.merge_cells("B5:F5")
            ws["B5"].value = "Госстандарта"
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(министерство, концерн, государственный комитет)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        title_first_sign()
        title_second_sign()
        
        ws.merge_cells("A12:O15")
        ws["A12"].value = "ПЕРЕЧЕНЬ МЕРОПРИЯТИЙ, НАПРАВЛЕННЫХ НА ДОСТИЖЕНИЕ ЦЕЛЕВЫХ ПОКАЗАТЕЛЕЙ В СФЕРЕ ЭНЕРГОСБЕРЕЖЕНИЯ ГОСУДАРСТВЕННОЙ ПРОГРАММЫ «УСТОЙЧИВАЯ ЭНЕРГЕТИКА И ЭНЕРГОЭФФЕКТИВНОСТЬ» "
        ws["A12"].font = bold_font_13
        ws["A12"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{plan.ministry.name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(министерство, концерн, государственный комитет)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"на {plan.year} год".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Целевые показатели:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:H25")
        ws["E25"].value = "энергосбережения"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:H26")
        ws["E26"].value = "по экономии ТЭР"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "по доле местных ТЭР в КПТ"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws.merge_cells("E28:H28")
        ws["E28"].value = "по доле ВИЭ в КПТ"
        ws["E28"].font = bold_font_11
        ws["E28"].alignment = left
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    
        ws["I28"].value = "-"
        ws["I28"].font = bold_font_11
        ws["I28"].alignment = center

        ws["J25"].value = f"{plan.energy_saving}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = center        
        ws["J26"].value = f"{plan.share_fuel}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = center        
        ws["J27"].value = f"{plan.saving_fuel}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = center        
        ws["J28"].value = f"{plan.share_energy}"
        ws["J28"].font = bold_font_11
        ws["J28"].alignment = center       

        ws["K25"].value = "%"
        ws["K25"].font = bold_font_11
        ws["K25"].alignment = center        
        ws["K26"].value = "т у.т."
        ws["K26"].font = bold_font_11
        ws["K26"].alignment = center        
        ws["K27"].value = "%"
        ws["K27"].font = bold_font_11
        ws["K27"].alignment = center        
        ws["K28"].value = "%"
        ws["K28"].font = bold_font_11
        ws["K28"].alignment = center
                     
        page_settings(ws, print_area = "A1:O32")
        
        return ws
 
    def reg_title_xlsx(wb, plan):
        ws = wb.create_sheet("Титульный лист", 0)
        columns = [("A", 8.43), ("B", 8.43), ("C", 8.43), ("D", 8.43),
                ("E", 8.43), ("F", 8.43), ("G", 8.43), ("H", 8.43),
                ("I", 8.43), ("J", 8.43), ("K", 8.43), ("L", 8.43),
                ("M", 8.43), ("N", 8.43)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 3:
                ws.row_dimensions[row].height = 12
            elif row == 5:
                ws.row_dimensions[row].height = 32.25
            elif row == 7:
                ws.row_dimensions[row].height = 4
            elif row == 8:
                ws.row_dimensions[row].height = 12
            elif row == 9:
                ws.row_dimensions[row].height = 22.5
            elif row == 16:
                ws.row_dimensions[row].height = 17.5
            elif row == 17:
                ws.row_dimensions[row].height = 19.5
            elif row == 20:
                ws.row_dimensions[row].height = 16.5
            else:
                ws.row_dimensions[row].height = 15
        
        def title_first_sign():
            ws.merge_cells("B1:D1")
            ws["B1"].value = "Согласовано".upper()
            ws["B1"].font = bold_font_11

            ws.merge_cells("B2:D2")
            ws["B2"].value = "_______________________"
            ws["B2"].font = bold_font_11
            
            ws.merge_cells("B3:D3")
            ws["B3"].value = "(должность)"
            ws["B3"].font = regular_font_9
            ws["B3"].alignment = center
                    
            ws.merge_cells("B4:F4")
            ws["B4"].value = "Департамента по энергоэффективности"
            ws["B4"].font = regular_font_11
            ws["B4"].alignment = left
                            
            ws.merge_cells("B5:F5")
            ws["B5"].value = "Госстандарта"
            ws["B5"].font = regular_font_11
            ws["B5"].alignment = left
                    
            ws.merge_cells("B6:D6")
            ws["B6"].value = "подписано ЭЦП"
            ws["B6"].font = regular_font_9_italic
            ws["B6"].alignment = center
            
            ws.merge_cells("B7:D7")
            ws["B7"].value = "_______________________"
            ws["B7"].font = bold_font_11
            
            ws.merge_cells("B8:D8")
            ws["B8"].value = "(подпись, инициалы и фамилия)"
            ws["B8"].font = regular_font_9
            ws["B8"].alignment = left
            
            ws.merge_cells("B9:E9")
            ws["B9"].value = "«___» ____________ 20__ г."
            ws["B9"].font = regular_font_11
            ws["B9"].alignment = left

        def title_second_sign():
            ws.merge_cells("K1:M1")
            ws["K1"].value = "Утверждаю".upper()
            ws["K1"].font = bold_font_11

            ws.merge_cells("K2:M2")
            ws["K2"].value = "_______________________"
            ws["K2"].font = bold_font_11
            
            ws.merge_cells("K3:M3")
            ws["K3"].value = "(должность)"
            ws["K3"].font = regular_font_9
            ws["K3"].alignment = center
                    
            ws.merge_cells("K4:M4")
            ws["K4"].value = "_______________________"
            ws["K4"].font = regular_font_11
            ws["K4"].alignment = left
                            
            ws.merge_cells("K5:M5")
            ws["K5"].value = "(обисполком, горисполком)"
            ws["K5"].font = regular_font_11
            ws["K5"].alignment = left
                    
            ws.merge_cells("K6:M6")
            ws["K6"].value = "подписано ЭЦП"
            ws["K6"].font = regular_font_9_italic
            ws["K6"].alignment = center
            
            ws.merge_cells("K7:M7")
            ws["K7"].value = "_______________________"
            ws["K7"].font = bold_font_11
            
            ws.merge_cells("K8:M8")
            ws["K8"].value = "(подпись, инициалы и фамилия)"
            ws["K8"].font = regular_font_9
            ws["K8"].alignment = left
            
            ws.merge_cells("K9:N9")
            ws["K9"].value = "«___» ____________ 20__ г."
            ws["K9"].font = regular_font_11
            ws["K9"].alignment = left
        
        title_first_sign()
        title_second_sign()
        
        ws.merge_cells("A12:O15")
        ws["A12"].value = "ПЕРЕЧЕНЬ МЕРОПРИЯТИЙ, НАПРАВЛЕННЫХ НА ДОСТИЖЕНИЕ ЦЕЛЕВЫХ ПОКАЗАТЕЛЕЙ В СФЕРЕ ЭНЕРГОСБЕРЕЖЕНИЯ ГОСУДАРСТВЕННОЙ ПРОГРАММЫ «УСТОЙЧИВАЯ ЭНЕРГЕТИКА И ЭНЕРГОЭФФЕКТИВНОСТЬ» "
        ws["A12"].font = bold_font_13
        ws["A12"].alignment = center
                                     
        ws.merge_cells("B16:N17")
        ws["B16"].value = f"{plan.region.name}"
        ws["B16"].font = regular_font_13
        ws["B16"].alignment = center
        
        for col in range(2, 15):
            ws.cell(row=17, column=col).border = bottom_border
        
        ws.merge_cells("D18:L18")
        ws["D18"].value = "(обисполком, горисполком)"
        ws["D18"].font = regular_font_13
        ws["D18"].alignment = center     
                
        ws.merge_cells("B20:N20")
        ws["B20"].value = f"на {plan.year} год".upper()
        ws["B20"].font = bold_font_13
        ws["B20"].alignment = center
        
        ws.merge_cells("B25:D25")
        ws["B25"].value = "Целевые показатели:"
        ws["B25"].font = bold_font_11
        ws["B25"].alignment = center    
        
        ws.merge_cells("E25:H25")
        ws["E25"].value = "энергосбережения"
        ws["E25"].font = bold_font_11
        ws["E25"].alignment = left 
               
        ws.merge_cells("E26:H26")
        ws["E26"].value = "по экономии ТЭР"
        ws["E26"].font = bold_font_11
        ws["E26"].alignment = left   
             
        ws.merge_cells("E27:H27")
        ws["E27"].value = "по доле местных ТЭР в КПТ"
        ws["E27"].font = bold_font_11
        ws["E27"].alignment = left 
                    
        ws.merge_cells("E28:H28")
        ws["E28"].value = "по доле ВИЭ в КПТ"
        ws["E28"].font = bold_font_11
        ws["E28"].alignment = left
                    
        ws["I25"].value = "-"
        ws["I25"].font = bold_font_11
        ws["I25"].alignment = center                    
        ws["I26"].value = "-"
        ws["I26"].font = bold_font_11
        ws["I26"].alignment = center                    
        ws["I27"].value = "-"
        ws["I27"].font = bold_font_11
        ws["I27"].alignment = center                    
        ws["I28"].value = "-"
        ws["I28"].font = bold_font_11
        ws["I28"].alignment = center

        ws["J25"].value = f"{plan.energy_saving}"
        ws["J25"].font = bold_font_11
        ws["J25"].alignment = center        
        ws["J26"].value = f"{plan.share_fuel}"
        ws["J26"].font = bold_font_11
        ws["J26"].alignment = center        
        ws["J27"].value = f"{plan.saving_fuel}"
        ws["J27"].font = bold_font_11
        ws["J27"].alignment = center        
        ws["J28"].value = f"{plan.share_energy}"
        ws["J28"].font = bold_font_11
        ws["J28"].alignment = center       

        ws["K25"].value = "%"
        ws["K25"].font = bold_font_11
        ws["K25"].alignment = center        
        ws["K26"].value = "т у.т."
        ws["K26"].font = bold_font_11
        ws["K26"].alignment = center        
        ws["K27"].value = "%"
        ws["K27"].font = bold_font_11
        ws["K27"].alignment = center        
        ws["K28"].value = "%"
        ws["K28"].font = bold_font_11
        ws["K28"].alignment = center
                     
        page_settings(ws, print_area = "A1:O32")
        
        return ws
 
    def first_half_xlsx(wb, plan):
        ws = wb.create_sheet("Часть 1")
        columns = [("A", 5.43), ("B", 58), ("C", 14), ("D", 10),
                ("E", 10), ("F", 10), ("G", 18.29)]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21
            elif row == 2:
                ws.row_dimensions[row].height = 12
            elif row == 3:
                ws.row_dimensions[row].height = 85.5
            elif row == 4:
                ws.row_dimensions[row].height = 34
            else:
                ws.row_dimensions[row].height = 15
        
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        ws.merge_cells("A1:G1")
        ws["A1"].value = "Часть 1. Показатели использования топливно-энергетических ресурсов"
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center
        ws.merge_cells("A2:G2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        configs = sorted(plan.column_configs, key=lambda x: x.year)
        headers = [
            "№ п/п", 
            "Основные показатели использования ТЭР", 
            "Единица измерения", 
            f"{plan.year - 2} г. {configs[0].label}", 
            f"{plan.year - 1} г. {configs[1].label}", 
            f"{plan.year} г. {configs[2].label}", 
            "Изменение ТЭР прогнозного года к предыдущему (увеличение +, уменьшение -), т у.т."
        ]
        
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = bold_font_11
            cell.alignment = center
            cell.border = thin_border
        
        def format_value(value, group):
            if value is None or value == '':
                return ''
            
            try:
                num = float(value)
            except (ValueError, TypeError):
                return value
            
            if group == 5:
                return round(num, 1)
            elif group in [6, 7, 8]:
                return round(num, 2)
            else:
                return int(round(num))
        
        previous_group = None
        row_index = 3

        for usage in sorted(plan.indicators_usage, 
            key=lambda u: (
                (0, u.indicator.Group) if u.indicator.Group is not None else (1, float('-inf')),
                (0, u.indicator.RowN) if u.indicator.RowN is not None else (1, float('-inf'))
            )
        ):
            group = usage.indicator.Group
            
            if group != previous_group:
                if group is not None:
                    try:
                        group_float = float(group)
                        if group_float.is_integer():
                            group_value = int(group_float)
                        else:
                            group_value = group
                    except (ValueError, TypeError):
                        group_value = group
                else:
                    group_value = ""
            else:
                group_value = ""
            
            previous_group = group
            
            if group in [5, 6]:
                QYearBeforePrev = 'x'
                QYearPrev = 'x'
                QYearCurrent = format_value(usage.QYearCurrent, group)
                difference = 'x'
            else:
                QYearBeforePrev = format_value(usage.QYearBeforePrev, group)
                QYearPrev = format_value(usage.QYearPrev, group)
                QYearCurrent = format_value(usage.QYearCurrent, group)
                difference = format_value((usage.QYearCurrent or 0) - (usage.QYearPrev or 0), group)
            
            row = [
                group_value,
                (usage.indicator.name if usage.indicator.name else "-") + 
                (f" ({usage.note})" if usage.note else ""),
                usage.indicator.unit.name,
                QYearBeforePrev,
                QYearPrev,
                QYearCurrent,
                difference,
            ]
            ws.append(row)
            
            row_index += 1
            
            if group_value in [5, 8]:
                ws.row_dimensions[row_index].height = 34
            
            for col in range(1, len(row) + 1):
                cell = ws.cell(row=row_index, column=col)
                
                if group_value != "":
                    cell.font = bold_font_11
                else:
                    cell.font = regular_font_11
                
                cell.alignment = center if col != 2 else left
                cell.border = thin_border
                
                if col == 1:
                    if group_value != "":
                        if isinstance(group_value, (int, float)):
                            try:
                                num = float(group_value)
                                if num.is_integer():
                                    cell.number_format = '0'
                                else:
                                    cell.number_format = '0.0'
                            except:
                                cell.number_format = '@'
                        else:
                            cell.number_format = '@'
                    else:
                        cell.number_format = '@'
                elif col in [4, 5, 6, 7]:
                    value = row[col-1]
                    if value != 'x' and value != '':
                        if group == 5:
                            cell.number_format = '0.0'
                        elif group in [6, 7, 8]:
                            cell.number_format = '0.00'
                        else:
                            cell.number_format = '0'
                    else:
                        cell.number_format = '@'
                else:
                    cell.number_format = '@'
        
        def org_small_signatures_indicators_xlsx(ws, start_row, plan):
            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"_______________ {org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"
            
            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=2, 
                    text=org_text, font=regular_font_10)
            
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+2, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+3, col_start=5, col_end=7, 
                    text="(юридическое лицо)", merge_direction='horizontal', font=regular_font_9, alignment=center)      
                     
            set_cell(ws, row_start=start_row+4, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+5, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+6, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+7, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                         
            set_cell(ws, row_start=start_row+9, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')               
                         
            set_cell(ws, row_start=start_row+10, col_start=5, col_end=7, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font=regular_font_9, row_height=11)  
                                 
            set_cell(ws, row_start=start_row+11, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+12, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+13, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+14, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                 
        def org_large_signatures_indicators_xlsx(ws, start_row, plan):
            set_cell(ws, row_start=start_row+2, col_start=2, col_end=2, merge_direction='horizontal', row_height=31,
                    text="Департамент по энергоэффективности Госстандарта", font=bold_font_11)
            
            set_cell(ws, row_start=start_row+3, col_start=2, row_end=start_row+4, col_end=2, 
                    text="Отдел анализа и прогнозирования развития энергосбережения производственно-технического управления", font=regular_font_10)
            
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+8, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')
            
            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"{org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"
            
            set_cell(ws, row_start=start_row+10, col_start=2, row_end=start_row+11, col_end=2, 
                    text=org_text, font=regular_font_10)
            
            set_cell(ws, row_start=start_row+12, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+13, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+14, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+15, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')   
                    
            set_cell(ws, row_start=start_row+3, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+4, col_start=5, col_end=7, 
                    text="(юридическое лицо)", merge_direction='horizontal', font=regular_font_9, alignment=center)      
                     
            set_cell(ws, row_start=start_row+5, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+6, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+7, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+8, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
                         
            set_cell(ws, row_start=start_row+10, col_start=5, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')               
                         
            set_cell(ws, row_start=start_row+11, col_start=5, col_end=7, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font=regular_font_9, row_height=11)  
                                 
            set_cell(ws, row_start=start_row+12, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+13, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+14, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+15, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')      
            
            set_cell(ws, row_start=start_row+17, col_start=5, col_end=8, 
                    text="_______________ облисполком (горисполком)", merge_direction='horizontal')               
                               
            set_cell(ws, row_start=start_row+18, col_start=5, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
                     
            set_cell(ws, row_start=start_row+19, col_start=5, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+20, col_start=5, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+21, col_start=5, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')     
                    
        def min_signatures_indicators_xlsx(ws, start_row, plan):
             
            set_cell(ws, row_start=start_row+2, col_start=2, col_end=2, merge_direction='horizontal', row_height=31,
                    text="Департамент по энергоэффективности Госстандарта", font=bold_font_11)
            
            set_cell(ws, row_start=start_row+3, col_start=2, row_end=start_row+4, col_end=2, 
                    text="Отдел анализа и прогнозирования развития энергосбережения производственно-технического управления", font=regular_font_10)
            
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+8, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+3, col_start=4, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+4, col_start=4, col_end=7, 
                    text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font=regular_font_9, alignment=left)      
                     
            set_cell(ws, row_start=start_row+5, col_start=4, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=left)
                     
            set_cell(ws, row_start=start_row+6, col_start=4, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+7, col_start=4, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+8, col_start=4, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')           
        
        def reg_signatures_indicators_xlsx(ws, start_row, plan):          
            set_cell(ws, row_start=start_row+2, col_start=2, col_end=2, merge_direction='horizontal', row_height=31,
                    text="Департамент по энергоэффективности Госстандарта", font=bold_font_11)
            
            set_cell(ws, row_start=start_row+3, col_start=2, row_end=start_row+4, col_end=2, 
                    text="Отдел анализа и прогнозирования развития энергосбережения производственно-технического управления", font=regular_font_10)
            
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+8, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                       
            set_cell(ws, row_start=start_row+10, col_start=2, col_end=2, 
                    text=f"{plan.region.name} областное (городское) управление по надзору за рациональным использованием ТЭР", merge_direction='horizontal')   
              
            set_cell(ws, row_start=start_row+11, col_start=2, col_end=2, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=bottom_left)   
              
            set_cell(ws, row_start=start_row+12, col_start=2, col_end=2, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+13, col_start=2, col_end=2, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=bottom_left) 
             
            set_cell(ws, row_start=start_row+14, col_start=2, col_end=2, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')  
                    
            set_cell(ws, row_start=start_row+3, col_start=4, col_end=7, 
                    text="__________________________________", merge_direction='horizontal')           
                                   
            set_cell(ws, row_start=start_row+4, col_start=4, col_end=7, 
                    text="(облисполком, горисполком)", merge_direction='horizontal', font=regular_font_9, alignment=left)      
                     
            set_cell(ws, row_start=start_row+5, col_start=4, col_end=7, 
                    text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=left)
                     
            set_cell(ws, row_start=start_row+6, col_start=4, col_end=7, 
                    text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)   
                       
            set_cell(ws, row_start=start_row+7, col_start=4, col_end=7, 
                    text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left) 
                         
            set_cell(ws, row_start=start_row+8, col_start=4, col_end=7, 
                    text="«___» ____________ 20__ г.", merge_direction='horizontal')   
        
        if export_type == "org_small":
            org_small_signatures_indicators_xlsx(ws, row_index + 1, plan) 
        elif export_type == "org_large":
            org_large_signatures_indicators_xlsx(ws, row_index + 1, plan)
        elif export_type == "ministry":
            min_signatures_indicators_xlsx(ws, row_index + 1, plan)
        elif export_type == "region":
            reg_signatures_indicators_xlsx(ws, row_index + 1, plan)
        else:
            ValueError("Erorr export_type 1sthalf")
            
        page_settings(ws, print_area = "A1:G75")
        return ws

    def create_events_sheet(wb, plan, sheet_name, title, is_increase=False):
        ws = wb.create_sheet(sheet_name)
        
        columns = [
            ("A", 4), ("B", 9), ("C", 15), ("D", 6),
            ("E", 6.75), ("F", 6.43), ("G", 6.86), ("H", 6.43),
            ("I", 9.43), ("J", 5), ("K", 7.86), ("L", 7.43),
            ("M", 6.57), ("N", 6.57), ("O", 6), ("P", 7.14),
            ("Q", 6.71), ("R", 7), ("S", 7.43)
        ]
        
        for col, width in columns:
            ws.column_dimensions[col].width = width

        for row in range(1, 34):
            if row == 1:
                ws.row_dimensions[row].height = 21.75
            elif row == 4:
                ws.row_dimensions[row].height = 23.25
            elif row == 5:
                ws.row_dimensions[row].height = 19.5
            elif row == 6:
                ws.row_dimensions[row].height = 159
            else:
                ws.row_dimensions[row].height = 15

        ws.merge_cells("A1:S1")
        ws["A1"].value = title
        ws["A1"].font = bold_font_13
        ws["A1"].alignment = center

        ws.merge_cells("A2:S2")
        ws["A2"].value = ""
        ws["A2"].font = bold_font_13
        ws["A2"].alignment = center

        ws.merge_cells("A3:A6"); ws["A3"].value = "№ п/п"
        ws.merge_cells("B3:B6"); ws["B3"].value = "Код основных направлений энергосбережения"
        ws.merge_cells("C3:C6"); ws["C3"].value = "Наименование мероприятия"
        ws.merge_cells("D3:D6"); ws["D3"].value = "Единицы измерения"
        ws.merge_cells("E3:E6"); ws["E3"].value = "Объем внедрения"
        ws.merge_cells("F3:G5") 
        if is_increase:
            ws["F3"].value = "Условно-годовое увеличение использо-вания местных ТЭР"
        else:
            ws["F3"].value = "Условно-годовой экономический эффект"
        ws["F6"].value = "т у.т."
        ws["G6"].value = "тыс. руб."
        ws.merge_cells("H3:H6"); ws["H3"].value = "Ожидаемый срок внедрения мероприятия, квартал"
        
        if is_increase:
            ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемое увеличение использования местных ТЭР от внедрения мероприятий в текущем году, т у.т."
        else:
            ws.merge_cells("I3:I6"); ws["I3"].value = "Ожидаемый экономический эффект от внедрения мероприятия в текущем году, т у.т."
        
        ws.merge_cells("J3:J6"); ws["J3"].value = "Срок окупаемости, лет"
        ws.merge_cells("K3:K6"); ws["K3"].value = "Общий объем финансирования, руб."
        ws.merge_cells("L3:L6"); ws["L3"].value = "Объем финансирования в текущем году, руб."
        ws.merge_cells("M3:S3"); ws["M3"].value = "источники финансирования, руб."
        ws.merge_cells("M4:P4"); ws["M4"].value = "бюджетные"
        ws.merge_cells("M5:M6"); ws["M5"].value = "республиканский бюджет на финансирование госпрограммы"
        ws.merge_cells("N5:N6"); ws["N5"].value = "республиканский бюджет"
        ws.merge_cells("O5:O6"); ws["O5"].value = "местный бюджет"
        ws.merge_cells("P5:P6"); ws["P5"].value = "другие"
        ws.merge_cells("Q4:Q6"); ws["Q4"].value = "собственные средства организации"
        ws.merge_cells("R4:R6"); ws["R4"].value = "кредиты банков, займы"
        ws.merge_cells("S4:S6"); ws["S4"].value = "иные"

        for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=19):
            for cell in row:
                if not cell.value:
                    continue
                if cell.coordinate in ("D3", "E3", "F6", "G6", "J3", "K3", "L3", "H3", "I3", "M5", "N5", "O5", "P5", "Q4", "R4", "S4"):
                    cell.alignment = vertical_text
                else:
                    cell.alignment = top
                cell.font = regular_font_10

        row_index = 7
        for col in range(1, 20):
            cell = ws.cell(row=row_index, column=col, value=col)
            cell.alignment = center
            cell.font = regular_font_10
            
        PERIOD_CODES = {"0001", "0002", "0003", "0004"}
        
        if is_increase:
            events = Event.query.filter(
                Event.id_plan == plan.id,
                Event.is_increase == True,
                Event.is_corrected == False,
                Event.display_code.notin_(PERIOD_CODES)
            ).all()
            events_corrected = Event.query.filter(
                Event.id_plan == plan.id,
                Event.is_increase == True,
                Event.is_corrected == True,
                Event.display_code.notin_(PERIOD_CODES)
            ).all()
            events_periods = Event.query.filter(
                Event.id_plan == plan.id,
                Event.is_increase == True,
                Event.display_code.in_(PERIOD_CODES)
            ).all()
            section_title_1 = "3.1 Мероприятия по увеличению использования местных ТЭР (первоначальная редакция)"
            section_title_2 = "3.2 Мероприятия по увеличению использования местных ТЭР (включенные в перечень при внесении в него изменений)"
            total_title = "Всего по части 3, в том числе:"
        else:
            events = Event.query.filter(
                Event.id_plan == plan.id,
                Event.is_econom == True,
                Event.is_corrected == False,
                Event.display_code.notin_(PERIOD_CODES)
            ).all()
            events_corrected = Event.query.filter(
                Event.id_plan == plan.id,
                Event.is_econom == True,
                Event.is_corrected == True,
                Event.display_code.notin_(PERIOD_CODES)
            ).all()
            events_periods = Event.query.filter(
                Event.id_plan == plan.id,
                Event.is_econom == True,
                Event.display_code.in_(PERIOD_CODES)
            ).all()
            section_title_1 = "Раздел 2.1 Мероприятия по экономии ТЭР (первоначальная редакция)"
            section_title_2 = "Раздел 2.2 Мероприятия по экономии ТЭР (включенные в план при внесении в него изменений)"
            total_title = "Всего по части 2, в том числе:"
        
        def add_section(title, execs, start_number=1):
            nonlocal row_index
            
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=19)
            ws.cell(row=row_index, column=1, value=title).font = bold_font_10
            ws.cell(row=row_index, column=1).alignment = center
            
            sum_cols = [6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19]
            sums = {col: 0 for col in sum_cols}
            
            if execs:
                for idx, econ in enumerate(execs, start=start_number):
                    row_index += 1
            
                    unit_name = ""
                    try:
                        if econ.direction and econ.direction.unit and econ.direction.unit.name:
                            unit_name = econ.direction.unit.name
                    except Exception:
                        pass
            
                    row = [
                        idx,
                        econ.display_code if econ.display_code else "",
                        econ.name if hasattr(econ, "name") else "",
                        unit_name,
                        econ.Volume if hasattr(econ, "Volume") else 0,
                        econ.EffTut if hasattr(econ, "EffTut") else 0,
                        econ.EffRub if hasattr(econ, "EffRub") else 0,
                        econ.ExpectedQuarter if hasattr(econ, "ExpectedQuarter") else 0,
                        econ.EffCurrYear if hasattr(econ, "EffCurrYear") else 0,
                        econ.Payback if hasattr(econ, "Payback") else 0,
                        econ.ObchVolumeFin if hasattr(econ, "ObchVolumeFin") else 0,
                        econ.VolumeFinCurrentYear if hasattr(econ, "VolumeFinCurrentYear") else 0,
                        econ.BudgetState if hasattr(econ, "BudgetState") else 0,
                        econ.BudgetRep if hasattr(econ, "BudgetRep") else 0,
                        econ.BudgetLoc if hasattr(econ, "BudgetLoc") else 0,
                        econ.BudgetOther if hasattr(econ, "BudgetOther") else 0,
                        econ.MoneyOwn if hasattr(econ, "MoneyOwn") else 0,
                        econ.MoneyLoan if hasattr(econ, "MoneyLoan") else 0,
                        econ.MoneyOther if hasattr(econ, "MoneyOther") else 0
                    ]
                    for col in sum_cols:
                        try:
                            sums[col] += float(row[col-1])
                        except (TypeError, ValueError):
                            pass
                    ws.append(row)
                    for col_idx in range(1, 20):
                        cell = ws.cell(row=row_index, column=col_idx)
                        cell.alignment = left if col_idx == 3 else center
                        cell.font = regular_font_10
                        if col_idx == 6 or col_idx == 9:
                            cell.number_format = '0.00'
                        elif col_idx == 10:
                            cell.number_format = '0.0'
                        elif col_idx in [5, 7, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
                            cell.number_format = '0'
            else:
                row_index += 1
                empty_row = [''] * 19
                ws.append(empty_row)
                for col_idx in range(1, 20):
                    cell = ws.cell(row=row_index, column=col_idx)
                    cell.alignment = center
                    cell.font = regular_font_10
            
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value="Итого по разделу:").alignment = left
            ws.cell(row=row_index, column=1).font = regular_font_10_italic
            
            for col in sum_cols:
                cell = ws.cell(row=row_index, column=col, value=sums[col])
                cell.alignment = center
                cell.font = regular_font_10_italic
                if col == 6 or col == 9:
                    cell.number_format = '0.00'
                elif col == 7 or col == 10:
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '0'
            return start_number + len(execs)

        def add_totals_and_periods(ws, row_index, events, events_corrected, events_periods):
            all_events = events + events_corrected
            sum_cols = [6, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19]
            
            def get_value(event, col):
                attr_map = {6: 'EffTut', 7: 'EffRub', 9: 'EffCurrYear', 10: 'Payback', 11: 'ObchVolumeFin',
                          12: 'VolumeFinCurrentYear', 13: 'BudgetState', 14: 'BudgetRep', 15: 'BudgetLoc',
                          16: 'BudgetOther', 17: 'MoneyOwn', 18: 'MoneyLoan', 19: 'MoneyOther'}
                val = getattr(event, attr_map.get(col, ''), 0)
                return float(val) if val else 0
            
            row_index += 1
            ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws.cell(row=row_index, column=1, value=total_title).font = bold_font_10
            ws.cell(row=row_index, column=1).alignment = left
            
            for col in sum_cols:
                total = sum(get_value(e, col) for e in all_events)
                cell = ws.cell(row=row_index, column=col, value=total)
                cell.alignment = center
                cell.font = bold_font_10
                if col == 6 or col == 9:
                    cell.number_format = '0.00'
                elif col == 7:
                    cell.number_format = '0.0'
                else:
                    cell.number_format = '0'
            
            PERIOD_CODES_LIST = ["0001", "0002", "0003", "0004"]
            
            period_events = [e for e in events_periods if e.display_code in PERIOD_CODES_LIST]
            period_events.sort(key=lambda x: PERIOD_CODES_LIST.index(x.display_code))
            
            for event in period_events:
                row_index += 1
                
                ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=8)
                
                event_name = event.name if event.name else ""
                ws.cell(row=row_index, column=1, value=event_name).font = regular_font_10
                ws.cell(row=row_index, column=1).alignment = left
                
                eff_curr_year = float(event.EffCurrYear) if event.EffCurrYear else 0
                cell = ws.cell(row=row_index, column=9, value=eff_curr_year)
                cell.alignment = center
                cell.font = regular_font_10
                cell.number_format = '0.00'
                
                for col in [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
                    ws.cell(row=row_index, column=col, value="")
            
            return row_index

        next_number = add_section(section_title_1, events, 1)
        second_number = add_section(section_title_2, events_corrected, next_number)
        row_index = add_totals_and_periods(ws, row_index, events, events_corrected, events_periods)

        for row in ws.iter_rows(min_row=3, max_row=row_index, min_col=1, max_col=19):
            for cell in row:
                cell.border = thin_border

        def org_small_signatures_events(ws, start_row, plan):
            plan_usd_rate = plan.usd_rate
            plan_cost_per_toe_usd = plan.cost_per_toe_usd
            
            text = f"Стоимость 1 т у.т. принята равной {plan_cost_per_toe_usd} долларов, при курсе {plan_usd_rate} руб."
            set_cell(ws, row_start=start_row, col_start=2, col_end=9, text=text, font=regular_font_10_italic, row_height=None)

            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"_______________ {org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"

            set_cell(ws, row_start=start_row+2, col_start=2, row_end=start_row+3, col_end=9, text=org_text)
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, text="__________________________________", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, text="(юридическое лицо)", merge_direction='horizontal', font=regular_font_9, alignment=center)
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+9, col_start=12, col_end=17, text="__________________________________", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+10, col_start=12, col_end=17, text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font=regular_font_9, row_height=11)
            set_cell(ws, row_start=start_row+11, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+12, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+13, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+14, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')

        def org_large_signatures_events(ws, start_row, plan):
            big_space = " " * 10
            text = f"Стоимость 1 т у.т. принята равной{big_space}долларов, при курсе{big_space}руб."
            set_cell(ws, row_start=start_row, col_start=2, col_end=9, text=text, font=regular_font_10_italic, row_height=None)

            set_cell(ws, row_start=start_row+2, col_start=2, col_end=9, merge_direction='horizontal', font=bold_font_10, text='Департамент по энергоэффективности Госстандарта')
            set_cell(ws, row_start=start_row+3, col_start=2, col_end=9, merge_direction='horizontal', text='Производственно-техническое управление')
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')

            set_cell(ws, row_start=start_row+9, col_start=2, col_end=9, row_end=start_row+10, text='Управление экономики и финансов')
            set_cell(ws, row_start=start_row+11, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+12, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+13, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+14, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')

            okpo = plan.organization.okpo if plan.organization else None
            org_name = get_org_name_by_okpo(okpo)
            
            if org_name:
                org_text = f"{org_name} по надзору за рациональным использованием ТЭР"
            else:
                org_text = f"_______________ областное (городское) управление по надзору за рациональным использованием ТЭР"

            set_cell(ws, row_start=start_row+16, col_start=2, merge_direction='horizontal', col_end=9, text=org_text)
            set_cell(ws, row_start=start_row+17, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+18, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+19, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+20, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, text="__________________________________", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, text="(юридическое лицо)", merge_direction='horizontal', font=regular_font_9, alignment=center)
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+9, col_start=12, col_end=17, text="__________________________________", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+10, col_start=12, col_end=17, text="(министерство, концерн, государственный коммитет)", merge_direction='horizontal', font=regular_font_9, row_height=11)
            set_cell(ws, row_start=start_row+11, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+12, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+13, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+14, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            
            set_cell(ws, row_start=start_row+16, col_start=12, col_end=17, row_height=40, text="_______________ облисполком (горисполком)", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+17, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+18, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+19, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+20, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')

        def min_signatures_events(ws, start_row, plan):
            big_space = " " * 10
            text = f"Стоимость 1 т у.т. принята равной{big_space}долларов, при курсе{big_space}руб."
            set_cell(ws, row_start=start_row, col_start=2, col_end=9, text=text, font=regular_font_10_italic, row_height=None)

            set_cell(ws, row_start=start_row+2, col_start=2, col_end=9, merge_direction='horizontal', text="Департамент по энергоэффективности Госстандарта", font=bold_font_11)
            set_cell(ws, row_start=start_row+3, col_start=2, col_end=9, row_height=50, merge_direction='horizontal', text="Производственно-техническое управление")
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')

            set_cell(ws, row_start=start_row+9, col_start=2, col_end=4, text="Управление экономики и финансов", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+10, col_start=2, col_end=4, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+11, col_start=2, col_end=4, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+12, col_start=2, col_end=4, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+13, col_start=2, col_end=4, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            
            set_cell(ws, row_start=start_row+2, col_start=12, col_end=17, text="__________________________________", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=16, text="(министерство, концерн, государственный комитет)", merge_direction='horizontal', font=regular_font_9, alignment=center)
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')

        def reg_signatures_events(ws, start_row, plan):
            big_space = " " * 10
            text = f"Стоимость 1 т у.т. принята равной{big_space}долларов, при курсе{big_space}руб."
            set_cell(ws, row_start=start_row, col_start=2, col_end=9, text=text, font=regular_font_10_italic, row_height=None)

            set_cell(ws, row_start=start_row+2, col_start=2, col_end=9, merge_direction='horizontal', font=bold_font_11, text="Департамент по энергоэффективности Госстандарта")
            set_cell(ws, row_start=start_row+4, col_start=2, col_end=9, merge_direction='horizontal', text="Производственно-техническое управление")
            set_cell(ws, row_start=start_row+5, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+6, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+7, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+8, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')

            set_cell(ws, row_start=start_row+9, col_start=2, col_end=9, merge_direction='horizontal', text="Управление экономики и финансов")
            set_cell(ws, row_start=start_row+10, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+11, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+12, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+13, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')

            set_cell(ws, row_start=start_row+14, col_start=2, col_end=9, row_height=30, merge_direction='horizontal',
                     text=f"{plan.region.name} областное (городское) управление по надзору за рациональным использованием ТЭР")
            set_cell(ws, row_start=start_row+15, col_start=2, col_end=3, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+16, col_start=2, col_end=3, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+17, col_start=2, col_end=3, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=center)
            set_cell(ws, row_start=start_row+18, col_start=2, col_end=3, text="«___» ____________ 20__ г.", merge_direction='horizontal')
            
            set_cell(ws, row_start=start_row+3, col_start=12, col_end=17, text="__________________________________", merge_direction='horizontal')
            set_cell(ws, row_start=start_row+4, col_start=12, col_end=16, text="(облисполком, горисполком)", merge_direction='horizontal', font=regular_font_9, alignment=center)
            set_cell(ws, row_start=start_row+5, col_start=12, col_end=15, text="Подписано ЭЦП", merge_direction='horizontal', font=regular_font_9_italic, alignment=center)
            set_cell(ws, row_start=start_row+6, col_start=12, col_end=15, text="_______________________", merge_direction='horizontal', row_height=5, alignment=bottom_left)
            set_cell(ws, row_start=start_row+7, col_start=12, col_end=15, text="(подпись, инициалы и фамилия)", merge_direction='horizontal', font=regular_font_9, row_height=11, alignment=left)
            set_cell(ws, row_start=start_row+8, col_start=12, col_end=15, text="«___» ____________ 20__ г.", merge_direction='horizontal')

        if export_type == "org_small":
            org_small_signatures_events(ws, row_index + 1, plan)
        elif export_type == "org_large":
            org_large_signatures_events(ws, row_index + 1, plan)
        elif export_type == "ministry":
            min_signatures_events(ws, row_index + 1, plan)
        elif export_type == "region":
            reg_signatures_events(ws, row_index + 1, plan)
        else:
            ValueError("Erorr export_type events")
        
        page_settings(ws, print_area="A1:S32")
        
        return ws

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    if export_type == "org_small":
        title_sheet = org_small_title_xlsx(wb, plan)
        wb.active = title_sheet
    elif export_type == "org_large":
        title_sheet = org_large_title_xlsx(wb, plan)
        wb.active = title_sheet
    elif export_type == "ministry":
        title_sheet = min_title_xlsx(wb, plan)
        wb.active = title_sheet
    elif export_type == "region":
        title_sheet = reg_title_xlsx(wb, plan) 
        wb.active = title_sheet
    else:
        ValueError("Erorr export_type")

    first_half_xlsx(wb, plan)
    
    econom_events = [event for event in plan.events if event.is_econom == True]
    increase_events = [event for event in plan.events if event.is_increase == True]
                
    if econom_events:
        create_events_sheet(wb, plan, "Часть 2", "Часть 2. Мероприятия по экономии топливно-энергетических ресурсов", False)

    if increase_events:
        create_events_sheet(wb, plan, "Часть 3", "Часть 3. Мероприятия по увеличению использования местных топливно-энергетических ресурсов", True)

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    if export_type == "org_small":
        filename = f"{plan.organization.okpo}_{plan.year}_{plan.id}.xlsx"
    elif export_type == "org_large":
        filename = f"{plan.organization.okpo}_{plan.year}_{plan.id}.xlsx"
    elif export_type == "ministry":
        filename = f"{plan.year}_{plan.id}.xlsx"
    elif export_type == "region":
        filename = f"{plan.year}_{plan.id}.xlsx"
    else:
        ValueError("Erorr export_type filename")
   
    return (
        file_stream,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename
    )